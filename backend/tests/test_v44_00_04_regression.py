"""v44.00.04 regression tests.

Covers:
- SMTP per-tenant scoping (master, reseller inherit, reseller own via direct DB insert)
- /engines/{name}/cascade-preview
- /pin-approvals/export (CSV + XLSX)
- /analytics/reseller-stats
- /marketplace/trending (public)
- verify-license key_not_found regression
- blacklist tenant isolation regression
"""
from __future__ import annotations
import io
import os
import pathlib
import zipfile

import pytest
import requests
from dotenv import dotenv_values
from pymongo import MongoClient

# --- Config ------------------------------------------------------------------
_FRONT_ENV = dotenv_values(pathlib.Path("/app/frontend/.env"))
BASE_URL = (_FRONT_ENV.get("REACT_APP_BACKEND_URL") or "").rstrip("/")
_BACK_ENV = dotenv_values(pathlib.Path("/app/backend/.env"))
MONGO_URL = _BACK_ENV.get("MONGO_URL") or os.environ.get("MONGO_URL")
DB_NAME = _BACK_ENV.get("DB_NAME") or os.environ.get("DB_NAME")

MASTER_KEY = "MS-C02AB012652A4FE692D69676"
RESELLER_STARTER = "MS-TESTBAYI-STARTER-V4371"
RESELLER_PRO = "MS-TESTBAYI-PRO-V4371"

assert BASE_URL, "REACT_APP_BACKEND_URL missing"
assert MONGO_URL and DB_NAME, "MongoDB env missing"

TIMEOUT = 30

_mongo = MongoClient(MONGO_URL)
_db = _mongo[DB_NAME]


# --- Fixtures ---------------------------------------------------------------
@pytest.fixture(scope="module")
def cleanup_smtp():
    """Snapshot reseller SMTP doc + delete test artefacts on teardown."""
    yield
    # Cleanup any reseller-owned smtp docs we inserted during test
    _db.settings.delete_many(
        {"_key": "smtp", "owner_license_key": {"$in": [RESELLER_STARTER, RESELLER_PRO]}}
    )


# ============================================================================
# 1. SMTP per-tenant scoping
# ============================================================================
class TestSMTPTenantScope:
    def test_master_smtp_own_scope(self):
        r = requests.get(
            f"{BASE_URL}/api/settings/smtp",
            headers={"X-Master-Key": MASTER_KEY},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("tenant_scoped") is True
        assert data.get("inherited_from_master") is False, data

    def test_reseller_starter_inherits_master(self, cleanup_smtp):
        # Ensure reseller has NO own smtp doc
        _db.settings.delete_many({"_key": "smtp", "owner_license_key": RESELLER_STARTER})

        r = requests.get(
            f"{BASE_URL}/api/settings/smtp",
            headers={"X-Master-Key": RESELLER_STARTER},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("tenant_scoped") is True
        assert data.get("inherited_from_master") is True, data

    def test_reseller_write_direct_db_creates_separate_doc(self, cleanup_smtp):
        """Simulate reseller saving own SMTP by direct DB write (since PUT on
        master panel returns 403 BAYI_ON_MASTER_PANEL)."""
        # Snapshot master's doc BEFORE
        master_doc_before = _db.settings.find_one(
            {"_key": "smtp",
             "$or": [{"owner_license_key": ""}, {"owner_license_key": {"$exists": False}}]}
        )
        # Insert reseller's own SMTP
        _db.settings.update_one(
            {"_key": "smtp", "owner_license_key": RESELLER_STARTER},
            {"$set": {
                "_key": "smtp",
                "owner_license_key": RESELLER_STARTER,
                "enabled": True,
                "auto_mode": False,
                "host": "smtp.reseller-test.example.com",
                "port": 2525,
                "username": "reseller-user",
                "password": "reseller-pass",
                "from_addr": "no-reply@reseller-test.example.com",
                "use_tls": "starttls",
            }},
            upsert=True,
        )
        # Verify: separate document exists
        reseller_doc = _db.settings.find_one(
            {"_key": "smtp", "owner_license_key": RESELLER_STARTER}
        )
        assert reseller_doc is not None
        assert reseller_doc["host"] == "smtp.reseller-test.example.com"

        # Master doc still intact
        master_doc_after = _db.settings.find_one(
            {"_key": "smtp",
             "$or": [{"owner_license_key": ""}, {"owner_license_key": {"$exists": False}}]}
        )
        if master_doc_before:
            assert master_doc_after is not None
            assert master_doc_after.get("host") == master_doc_before.get("host")

        # GET as reseller: now shows OWN settings, inherited=false
        r = requests.get(
            f"{BASE_URL}/api/settings/smtp",
            headers={"X-Master-Key": RESELLER_STARTER},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["host"] == "smtp.reseller-test.example.com"
        assert data["port"] == 2525
        assert data["inherited_from_master"] is False


# ============================================================================
# 2. Engine cascade preview
# ============================================================================
class TestEngineCascadePreview:
    def test_master_cascade_preview(self):
        r = requests.get(
            f"{BASE_URL}/api/engines/spamassassin/cascade-preview",
            headers={"X-Master-Key": MASTER_KEY},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        for k in ("engine", "label", "current_state", "target_state",
                  "affected_total", "will_change", "already_same", "samples"):
            assert k in data, f"missing key {k}: {data}"
        assert data["engine"] == "spamassassin"
        assert isinstance(data["current_state"], bool)
        assert data["target_state"] == (not data["current_state"])
        assert isinstance(data["samples"], list)
        assert len(data["samples"]) <= 8
        # will_change + already_same = affected_total
        assert data["will_change"] + data["already_same"] == data["affected_total"], data
        # sample shape
        for s in data["samples"]:
            assert "customer_name" in s
            assert "plan" in s
            assert "license_key_short" in s
            assert "current_state" in s

    def test_cascade_preview_does_not_modify_db(self):
        # Count engines rows before/after
        before = _db.engines.count_documents({"name": "spamassassin"})
        for _ in range(3):
            requests.get(
                f"{BASE_URL}/api/engines/spamassassin/cascade-preview",
                headers={"X-Master-Key": MASTER_KEY},
                timeout=TIMEOUT,
            )
        after = _db.engines.count_documents({"name": "spamassassin"})
        # Preview may bootstrap master's own row via _engines_for(owner="") — this is
        # idempotent. Row count should not increase after first call.
        assert after >= before  # allow bootstrap
        # But state should not have flipped — grab enabled values snapshot
        # and compare across calls: they must be identical.
        snap1 = {d["owner_license_key"]: d.get("enabled") for d in
                 _db.engines.find({"name": "spamassassin"},
                                  {"_id": 0, "owner_license_key": 1, "enabled": 1})}
        requests.get(
            f"{BASE_URL}/api/engines/spamassassin/cascade-preview",
            headers={"X-Master-Key": MASTER_KEY},
            timeout=TIMEOUT,
        )
        snap2 = {d["owner_license_key"]: d.get("enabled") for d in
                 _db.engines.find({"name": "spamassassin"},
                                  {"_id": 0, "owner_license_key": 1, "enabled": 1})}
        assert snap1 == snap2, "cascade-preview mutated engine states!"

    def test_reseller_cascade_preview_forbidden(self):
        r = requests.get(
            f"{BASE_URL}/api/engines/spamassassin/cascade-preview",
            headers={"X-Master-Key": RESELLER_STARTER},
            timeout=TIMEOUT,
        )
        assert r.status_code == 403, r.text


# ============================================================================
# 3. PIN history export
# ============================================================================
class TestPinExport:
    def test_csv_export_master_via_query(self):
        r = requests.get(
            f"{BASE_URL}/api/pin-approvals/export",
            params={"fmt": "csv", "license_key": MASTER_KEY},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200, r.text
        assert "text/csv" in r.headers.get("content-type", "").lower()
        body = r.content
        assert body[:3] == b"\xef\xbb\xbf", f"missing UTF-8 BOM: {body[:20]!r}"
        # header row = 13 columns
        text = body.decode("utf-8-sig")
        header = text.splitlines()[0]
        cols = header.split(",")
        assert len(cols) == 13, f"expected 13 cols, got {len(cols)}: {cols}"
        expected = {"bayi_name", "bayi_email", "company", "plan", "license_key",
                    "requested_at", "decided_at", "status", "pin_length",
                    "pin_hash_preview", "requested_ip", "decided_by_ip", "note"}
        assert set(cols) == expected, cols

    def test_xlsx_export_master(self):
        r = requests.get(
            f"{BASE_URL}/api/pin-approvals/export",
            params={"fmt": "xlsx", "license_key": MASTER_KEY},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "").lower()
        assert "openxmlformats" in ct or "spreadsheetml" in ct, ct
        # Body must be a valid ZIP (xlsx = zipped xml)
        assert r.content[:2] == b"PK", "not a zip/xlsx binary"
        # Try to open as workbook
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(r.content), read_only=True)
            ws = wb.active
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            assert len(header) == 13, header
        except ImportError:
            # openpyxl not available in test env — zip check is sufficient
            with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
                names = zf.namelist()
            assert any("workbook" in n or "sheet" in n for n in names), names

    def test_export_no_key_forbidden(self):
        r = requests.get(f"{BASE_URL}/api/pin-approvals/export",
                         params={"fmt": "csv"}, timeout=TIMEOUT)
        assert r.status_code == 403, r.text

    def test_export_reseller_forbidden(self):
        r = requests.get(
            f"{BASE_URL}/api/pin-approvals/export",
            params={"fmt": "csv", "license_key": RESELLER_STARTER},
            timeout=TIMEOUT,
        )
        assert r.status_code == 403, r.text

    def test_export_status_filter_approved(self):
        r = requests.get(
            f"{BASE_URL}/api/pin-approvals/export",
            params={"fmt": "csv", "license_key": MASTER_KEY, "status": "approved"},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200
        text = r.content.decode("utf-8-sig")
        lines = text.splitlines()
        if len(lines) > 1:
            # Column position of status = index 7 (after 7 fields before)
            header = lines[0].split(",")
            status_idx = header.index("status")
            for row in lines[1:]:
                parts = row.split(",")
                if len(parts) > status_idx:
                    val = parts[status_idx].strip().strip('"')
                    assert val == "approved", f"row status not approved: {val}"


# ============================================================================
# 4. Reseller analytics
# ============================================================================
class TestResellerAnalytics:
    def test_master_scope(self):
        r = requests.get(
            f"{BASE_URL}/api/analytics/reseller-stats",
            headers={"X-Master-Key": MASTER_KEY},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("scope") == "master"
        assert "license_key_short" in data
        counts = data.get("counts", {})
        for k in ("scanned_24h", "blocked_24h", "scanned_7d", "blocked_7d",
                  "scanned_30d", "blocked_30d"):
            assert k in counts, counts
            assert isinstance(counts[k], int)
        q = data.get("quarantine", {})
        for k in ("total", "phish", "virus"):
            assert k in q
        assert isinstance(data.get("block_rate_pct"), (int, float))
        assert 0.0 <= float(data["block_rate_pct"]) <= 100.0
        assert isinstance(data.get("top_threats"), list)
        assert len(data["top_threats"]) <= 5
        assert isinstance(data.get("estimated_savings_usd_30d"), (int, float))
        trend = data.get("trend_7d")
        assert isinstance(trend, list) and len(trend) == 7, trend
        for t in trend:
            assert "day" in t and "blocked" in t
        engines = data.get("engines", {})
        assert "active" in engines and "total" in engines

    def test_reseller_scope(self):
        r = requests.get(
            f"{BASE_URL}/api/analytics/reseller-stats",
            headers={"X-Master-Key": RESELLER_STARTER},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("scope") == "reseller"
        assert data.get("license_key_short", "").startswith("MS-TESTBAYI")

    def test_missing_key_401(self):
        r = requests.get(f"{BASE_URL}/api/analytics/reseller-stats", timeout=TIMEOUT)
        assert r.status_code == 401, r.text


# ============================================================================
# 5. Marketplace trending (public)
# ============================================================================
class TestMarketplaceTrending:
    def test_trending_public_no_auth(self):
        r = requests.get(
            f"{BASE_URL}/api/marketplace/trending",
            params={"limit": 6, "days": 30},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("period_days") == 30
        assert "generated_at" in data
        assert isinstance(data.get("trending"), list)
        totals = data.get("totals", {})
        assert "active_signatures" in totals
        assert "total_installs_all_time" in totals

        for item in data["trending"]:
            for k in ("id", "name", "category", "target", "score",
                      "installs", "upvotes", "publisher_label"):
                assert k in item, f"missing {k}: {item}"
            # publisher_label must be anonymised (starts with '@', no full key)
            assert item["publisher_label"].startswith("@"), item["publisher_label"]
            assert "MS-" not in item["publisher_label"]

    def test_trending_no_headers_still_works(self):
        # explicitly no auth headers of any kind
        sess = requests.Session()
        sess.headers.clear()
        sess.headers["User-Agent"] = "trending-test"
        r = sess.get(f"{BASE_URL}/api/marketplace/trending", timeout=TIMEOUT)
        assert r.status_code == 200, r.text


# ============================================================================
# 6. Regression: verify-license key_not_found
# ============================================================================
class TestVerifyLicenseRegression:
    def test_bogus_key_404_and_violation(self):
        bogus = "MS-BOGUS-DOES-NOT-EXIST-V44004"
        r = requests.post(
            f"{BASE_URL}/api/plugin/verify-license",
            json={"license_key": bogus, "server_ip": "5.6.7.8"},
            timeout=TIMEOUT,
        )
        assert r.status_code == 404, r.text
        # Check violation row
        vio = _db.license_violations.find_one(
            {"license_key": bogus, "reason": "key_not_found"},
            sort=[("_id", -1)],
        )
        assert vio is not None, "key_not_found violation not recorded"


# ============================================================================
# 7. Regression: blacklist tenant isolation
# ============================================================================
class TestBlacklistTenantRegression:
    def test_reseller_sees_empty_or_own_only(self):
        r = requests.get(
            f"{BASE_URL}/api/blacklist/requests",
            headers={"X-Master-Key": RESELLER_STARTER},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200, r.text
        items = r.json() if isinstance(r.json(), list) else r.json().get("items", [])
        # All items must be owned by reseller
        for it in items:
            owner = it.get("owner_license_key", "")
            assert owner == RESELLER_STARTER, f"leak: {owner}"

    def test_master_sees_all(self):
        r = requests.get(
            f"{BASE_URL}/api/blacklist/requests",
            headers={"X-Master-Key": MASTER_KEY},
            timeout=TIMEOUT,
        )
        assert r.status_code == 200, r.text
