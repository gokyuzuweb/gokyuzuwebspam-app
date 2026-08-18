"""v43.89 — Gelişmiş Mail Raporlama (PDF + Excel export).

Endpoints:
- POST /api/reports/mail-activity  → JSON | PDF | Excel export
  Params: {email, direction: sent|received|both, days, format}

Kullanım örneği:
  x@x.com'un gönderdiği tüm mailler (son 30g) → alıcı listesi + verdict + tarihler
  x@x.com'a gelen tüm mailler → gönderen listesi + verdict + tarihler
"""
from __future__ import annotations
import os
import uuid
from datetime import datetime, timezone, timedelta
from io import BytesIO
from typing import Optional, Literal

from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import Response, JSONResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field

_client = AsyncIOMotorClient(os.environ["MONGO_URL"])
db = _client[os.environ["DB_NAME"]]

router = APIRouter(prefix="/reports", tags=["reports"])


def _iso() -> str:
    return datetime.now(timezone.utc).isoformat()


class ReportRequest(BaseModel):
    email: str = Field(..., min_length=3)
    direction: Literal["sent", "received", "both"] = "both"
    days: int = Field(30, ge=1, le=365)
    format: Literal["json", "pdf", "xlsx"] = "json"
    limit: int = Field(1000, ge=10, le=5000)


async def _collect_events(email: str, direction: str, days: int, limit: int) -> dict:
    """mail_events + quarantine koleksiyonlarından ilgili maili çek."""
    since = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    email_lc = email.strip().lower()
    sent, received = [], []

    # SENT: from_addr == email  → alıcı raporu
    if direction in ("sent", "both"):
        cursor = db.mail_events.find(
            {"from_addr": {"$regex": f"^{email_lc}$", "$options": "i"},
             "ts": {"$gte": since}},
            {"_id": 0, "ts": 1, "from_addr": 1, "to_addr": 1,
             "subject": 1, "verdict": 1, "sender_ip": 1, "engine": 1,
             "size_bytes": 1, "spam_score": 1},
        ).sort("ts", -1).limit(limit)
        sent = await cursor.to_list(limit)

    # RECEIVED: to_addr contains email
    if direction in ("received", "both"):
        cursor = db.mail_events.find(
            {"to_addr": {"$regex": email_lc, "$options": "i"},
             "ts": {"$gte": since}},
            {"_id": 0, "ts": 1, "from_addr": 1, "to_addr": 1,
             "subject": 1, "verdict": 1, "sender_ip": 1, "engine": 1,
             "size_bytes": 1, "spam_score": 1},
        ).sort("ts", -1).limit(limit)
        received = await cursor.to_list(limit)

    # Aggregate özet
    def _summ(rows: list) -> dict:
        by_verdict = {}
        peers = {}   # sent → alıcılar, received → göndericiler
        for r in rows:
            v = r.get("verdict") or "unknown"
            by_verdict[v] = by_verdict.get(v, 0) + 1
            peer = r.get("to_addr" if rows is sent else "from_addr") or ""
            if peer:
                peers[peer] = peers.get(peer, 0) + 1
        top_peers = sorted(peers.items(), key=lambda x: -x[1])[:20]
        return {"total": len(rows), "by_verdict": by_verdict,
                "top_peers": [{"peer": p, "count": c} for p, c in top_peers]}

    return {
        "email": email,
        "direction": direction,
        "days": days,
        "generated_at": _iso(),
        "sent": {"rows": sent, "summary": _summ(sent) if sent else {"total": 0, "by_verdict": {}, "top_peers": []}},
        "received": {"rows": received, "summary": _summ(received) if received else {"total": 0, "by_verdict": {}, "top_peers": []}},
    }


def _build_report_pdf(data: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor
    from reportlab.pdfgen import canvas
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors as _rc

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4

    # Header band
    c.setFillColor(HexColor("#0f172a"))
    c.rect(0, H - 3 * cm, W, 3 * cm, fill=True, stroke=False)
    c.setFillColor(HexColor("#a5b4fc"))
    c.setFont("Helvetica-Bold", 16)
    c.drawString(2 * cm, H - 1.5 * cm, "GökyüzüWebSpam · Mail Aktivite Raporu")
    c.setFillColor(HexColor("#94a3b8"))
    c.setFont("Helvetica", 9)
    c.drawString(2 * cm, H - 2.3 * cm,
                 f"Email: {data['email']} · Yön: {data['direction']} · Son {data['days']} gün · {data['generated_at'][:19]} UTC")

    y = H - 4 * cm

    def _draw_section(title: str, section: dict, y_start: float) -> float:
        c.setFillColor(HexColor("#e2e8f0"))
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2 * cm, y_start, title)
        c.setFillColor(HexColor("#94a3b8"))
        c.setFont("Helvetica", 9)
        s = section["summary"]
        c.drawString(2 * cm, y_start - 0.5 * cm,
                     f"Toplam: {s['total']} · Verdict: " +
                     ", ".join(f"{k}={v}" for k, v in s["by_verdict"].items()))
        # Top peers table
        y2 = y_start - 1.2 * cm
        rows = [["#", "Peer", "Adet"]]
        for i, p in enumerate(s["top_peers"][:12]):
            rows.append([str(i + 1), p["peer"][:44], str(p["count"])])
        if len(rows) > 1:
            tbl = Table(rows, colWidths=[1 * cm, 12 * cm, 2 * cm])
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1e293b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#a5b4fc")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.25, HexColor("#334155")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#0f172a"), HexColor("#111827")]),
                ("TEXTCOLOR", (0, 1), (-1, -1), _rc.white),
            ]))
            tbl_h = tbl.wrap(W - 4 * cm, 10 * cm)[1]
            tbl.drawOn(c, 2 * cm, y2 - tbl_h)
            return y2 - tbl_h - 0.5 * cm
        return y2 - 0.5 * cm

    if data["direction"] in ("sent", "both") and data["sent"]["summary"]["total"] > 0:
        y = _draw_section(f"📤 GÖNDERİLEN — {data['email']} → alıcılar", data["sent"], y)
    if data["direction"] in ("received", "both") and data["received"]["summary"]["total"] > 0:
        y = _draw_section(f"📥 GELEN — göndericiler → {data['email']}", data["received"], y)

    c.setFillColor(HexColor("#64748b"))
    c.setFont("Helvetica-Oblique", 7)
    c.drawString(2 * cm, 1 * cm, "GökyüzüWebSpam · Mail Activity Report v43.89")
    c.showPage()
    c.save()
    return buf.getvalue()


def _build_report_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Özet"
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="A5B4FC")

    ws_meta.append(["Alan", "Değer"])
    ws_meta.append(["Email", data["email"]])
    ws_meta.append(["Yön", data["direction"]])
    ws_meta.append(["Kapsam", f"Son {data['days']} gün"])
    ws_meta.append(["Oluşturma", data["generated_at"]])
    ws_meta.append([])
    ws_meta.append(["Yön", "Toplam", "Verdict Dağılımı"])
    for direction_key in ("sent", "received"):
        if data[direction_key]["summary"]["total"] > 0:
            s = data[direction_key]["summary"]
            vs = ", ".join(f"{k}={v}" for k, v in s["by_verdict"].items())
            ws_meta.append([direction_key, s["total"], vs])
    for row in ws_meta["A1:C1"]:
        for c in row:
            c.fill = header_fill
            c.font = header_font
    ws_meta.column_dimensions["A"].width = 16
    ws_meta.column_dimensions["B"].width = 30
    ws_meta.column_dimensions["C"].width = 50

    def _add_sheet(name: str, rows: list, peer_key: str) -> None:
        ws = wb.create_sheet(name)
        cols = ["Tarih (UTC)", "From", "To", "Konu", "Verdict", "Score", "Engine", "Sender IP", "Size (B)"]
        ws.append(cols)
        for c in ws[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
        widths = [22, 32, 32, 60, 14, 8, 12, 18, 10]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w
        for r in rows:
            ws.append([
                (r.get("ts") or "")[:19],
                r.get("from_addr", ""),
                r.get("to_addr", ""),
                (r.get("subject") or "")[:120],
                r.get("verdict", ""),
                r.get("spam_score") or "",
                r.get("engine", ""),
                r.get("sender_ip", ""),
                r.get("size_bytes") or "",
            ])
        ws.freeze_panes = "A2"

    if data["direction"] in ("sent", "both"):
        _add_sheet("Gönderilen", data["sent"]["rows"], "to_addr")
    if data["direction"] in ("received", "both"):
        _add_sheet("Gelen", data["received"]["rows"], "from_addr")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/mail-activity")
async def mail_activity(payload: ReportRequest, request: Request):
    """Mail aktivite raporu — JSON / PDF / XLSX formatlarında."""
    hdr = request.headers
    if not (hdr.get("X-Master-Key") or hdr.get("X-License-Key")):
        raise HTTPException(401, "Lisans anahtarı gerekli")

    data = await _collect_events(payload.email, payload.direction,
                                  payload.days, payload.limit)

    # Audit
    try:
        await db.audit_logs.insert_one({
            "id": str(uuid.uuid4()), "action": "mail_report_generated",
            "actor_ip": (hdr.get("x-forwarded-for") or "").split(",")[0].strip(),
            "details": {"email": payload.email, "direction": payload.direction,
                         "days": payload.days, "format": payload.format,
                         "sent_total": data["sent"]["summary"]["total"],
                         "received_total": data["received"]["summary"]["total"]},
            "at": _iso(), "severity": "info",
        })
    except Exception:
        pass

    fname_safe = payload.email.replace("@", "_at_").replace("/", "_")[:60]
    fname_base = f"mail-report-{fname_safe}-{payload.days}d"

    if payload.format == "pdf":
        pdf = _build_report_pdf(data)
        return Response(content=pdf, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="{fname_base}.pdf"'})
    if payload.format == "xlsx":
        xlsx = _build_report_xlsx(data)
        return Response(content=xlsx,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="{fname_base}.xlsx"'})
    # JSON default
    return data
